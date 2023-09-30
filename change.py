import openpyxl
# Создаем пустой список для хранения данных
data = []
# Открываем файл XLSX
workbook1 = openpyxl.load_workbook('saits.xlsx')

# Получаем активный лист (можно указать имя листа, если он не активен)
sheet = workbook1.active
# Проходимся по строкам листа и собираем данные в массив
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Обновляем значения в массиве, удаляя часть строки до первого символа "/"
workbook = openpyxl.Workbook()
worksheet = workbook.active
# Заголовки столбцов
worksheet['A1'] = 'URL'
for j, cell in enumerate(data, start=2):
        print(cell)
        cell = cell[0].replace("https://" , "").replace("http://" , "")
        print(cell)
        worksheet[f'A{j}'] = cell
workbook.save('change_saits.xlsx')

# Закройте книгу
workbook.close()
# Закрываем файл XLSX после использования
workbook1.close()