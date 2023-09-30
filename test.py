
import requests
import openpyxl

# Список URL-адресов, на которые нужно отправить GET-запросы
urls = []

# Открываем файл XLSX
workbook1 = openpyxl.load_workbook('change_saits.xlsx')

# Получаем активный лист (можно указать имя листа, если он не активен)
sheet = workbook1.active



# Проходимся по строкам листа и собираем данные в массив
for row in sheet.iter_rows(values_only=True):
    url = f'http://127.0.0.1:5000/check_domain?domain={row[0]}'
    urls.append(url)

# Создайте новую книгу Excel и активный лист
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Заголовки столбцов
worksheet['A1'] = 'URL'
worksheet['B1'] = 'Response'

# Отправьте GET-запросы и запишите ответы в таблицу
index=0
for  url in urls[200:]:
    response = requests.get(url)
    print(url)
    if (len(response.text)>10 and response.status_code==200):
        worksheet[f'A{index+1}'] = url
        worksheet[f'B{index+1}'] = response.text
        index+=1
    if index==500:
        break

# Сохраните книгу в файл XLSX
workbook.save('responses.xlsx')

# Закройте книгу
workbook.close()

print("Ответы записаны в файл responses.xlsx")

